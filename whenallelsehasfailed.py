import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
import numpy as np

np.set_printoptions(threshold=np.nan)


def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = np.genfromtxt(filename, skip_header = 1, delimiter=[30,16,16,16,16,16,16,16,16,16,16,16])  
    print(book[85])
    print("hello")
    """index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in xrange(0, nrows):
        for col in xrange(0, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)
    """
    return book

open_xls_as_xlsx("/Users/lucychen/Documents/Lab/NIH lab/Data/exportedTestTxt.txt")
